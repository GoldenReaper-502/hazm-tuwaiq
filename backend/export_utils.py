"""Export utilities for reports to PDF and Excel formats.

Requires reportlab for PDF and openpyxl for Excel (optional).
"""
from typing import Any, Dict, List, Optional
import json
from datetime import datetime


def export_to_json(reports: List[Dict[str, Any]]) -> str:
    """Export reports as JSON string."""
    return json.dumps(reports, indent=2)


def export_to_csv(reports: List[Dict[str, Any]]) -> str:
    """Export reports as CSV (simple, no complex nesting)."""
    import csv
    import io
    
    output = io.StringIO()
    fieldnames = [
        "id", "detection_id", "timestamp", "location", "risk_score", 
        "risk_level", "objects_count", "summary"
    ]
    
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    
    for report in reports:
        writer.writerow({
            "id": report.get("id", ""),
            "detection_id": report.get("detection_id", ""),
            "timestamp": report.get("timestamp", ""),
            "location": report.get("location", ""),
            "risk_score": report.get("risk_score", ""),
            "risk_level": report.get("risk_level", ""),
            "objects_count": report.get("objects_count", ""),
            "summary": report.get("summary", "").replace("\n", "; ")[:100],
        })
    return output.getvalue()


def export_to_pdf(reports: List[Dict[str, Any]], filename: Optional[str] = None) -> Optional[bytes]:
    """Export reports as PDF.
    
    Returns PDF bytes if reportlab available, else None.
    """
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib import colors
        from io import BytesIO
    except ImportError:
        return None
    
    buffer = BytesIO()
    if not filename:
        filename = f"reports_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#1a1a1a"),
        spaceAfter=12,
    )
    
    risk_style_high = ParagraphStyle(
        "RiskHigh",
        parent=styles["Normal"],
        textColor=colors.red,
        fontSize=12,
        fontName="Helvetica-Bold",
    )
    risk_style_medium = ParagraphStyle(
        "RiskMedium",
        parent=styles["Normal"],
        textColor=colors.orange,
        fontSize=12,
        fontName="Helvetica-Bold",
    )
    risk_style_low = ParagraphStyle(
        "RiskLow",
        parent=styles["Normal"],
        textColor=colors.green,
        fontSize=12,
        fontName="Helvetica-Bold",
    )
    
    # Title
    story.append(Paragraph("Safety Detection Reports", title_style))
    story.append(Spacer(1, 0.2*inch))
    story.append(Paragraph(f"Generated: {datetime.utcnow().isoformat()}", styles["Normal"]))
    story.append(Spacer(1, 0.3*inch))
    
    # Reports
    for report in reports:
        risk_level = report.get("risk_level", "UNKNOWN")
        if risk_level == "HIGH":
            risk_style = risk_style_high
        elif risk_level == "MEDIUM":
            risk_style = risk_style_medium
        else:
            risk_style = risk_style_low
        
        # Report header
        story.append(Paragraph(f"Report ID: {report.get('id')}", styles["Heading2"]))
        story.append(Paragraph(
            f"Risk Level: {risk_level} (Score: {report.get('risk_score')})",
            risk_style
        ))
        story.append(Paragraph(f"Location: {report.get('location')}", styles["Normal"]))
        story.append(Paragraph(f"Timestamp: {report.get('timestamp')}", styles["Normal"]))
        story.append(Spacer(1, 0.1*inch))
        
        # Summary
        summary_text = report.get("summary", "").replace("\n", "<br/>")
        story.append(Paragraph(f"<b>Summary:</b> {summary_text}", styles["Normal"]))
        story.append(Spacer(1, 0.1*inch))
        
        # Recommendations
        recs = report.get("recommendations", [])
        if recs:
            story.append(Paragraph("<b>Recommendations:</b>", styles["Normal"]))
            for rec in recs:
                story.append(Paragraph(f"• {rec}", styles["Normal"]))
        
        story.append(Spacer(1, 0.3*inch))
        story.append(PageBreak())
    
    doc.build(story)
    return buffer.getvalue()


def export_to_excel(reports: List[Dict[str, Any]], filename: Optional[str] = None) -> Optional[bytes]:
    """Export reports as Excel (.xlsx).
    
    Returns Excel bytes if openpyxl available, else None.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
    except ImportError:
        return None
    
    buffer = BytesIO()
    if not filename:
        filename = f"reports_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reports"
    
    # Headers
    headers = ["ID", "Detection ID", "Timestamp", "Location", "Risk Level", "Risk Score", "Objects", "Summary"]
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for report in reports:
        risk_level = report.get("risk_level", "UNKNOWN")
        risk_color = {"HIGH": "FFC7CE", "MEDIUM": "FFEB9C", "LOW": "C6EFCE"}.get(risk_level, "FFFFFF")
        risk_font_color = {"HIGH": "9C0006", "MEDIUM": "9C6500", "LOW": "006100"}.get(risk_level, "000000")
        
        row_data = [
            report.get("id"),
            report.get("detection_id"),
            report.get("timestamp"),
            report.get("location"),
            risk_level,
            report.get("risk_score"),
            report.get("objects_count"),
            report.get("summary", "").replace("\n", "; ")[:50] + "...",
        ]
        ws.append(row_data)
        
        # Style row
        row_num = ws.max_row
        for cell in ws[row_num]:
            cell.fill = PatternFill(start_color=risk_color, end_color=risk_color, fill_type="solid")
            cell.font = Font(color=risk_font_color)
            cell.alignment = Alignment(horizontal="left", wrap_text=True)
    
    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 50
    
    wb.save(buffer)
    return buffer.getvalue()
