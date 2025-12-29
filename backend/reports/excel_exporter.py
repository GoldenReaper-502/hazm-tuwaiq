"""
HAZM TUWAIQ - Excel Report Exporter
Professional Excel reports with formatting and charts using openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.chart import (
    BarChart, LineChart, PieChart, Reference
)
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict, Any, Optional

from .models import ReportData, ReportMetrics, ReportTable, ReportChart


class ExcelReportExporter:
    """Generate professional Excel reports"""
    
    def __init__(self):
        self.wb = None
        self._setup_styles()
    
    def _setup_styles(self):
        """Define Excel styles"""
        self.header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
        
        self.title_font = Font(name='Arial', size=18, bold=True, color='1E293B')
        self.subtitle_font = Font(name='Arial', size=12, bold=True, color='475569')
        self.normal_font = Font(name='Arial', size=10, color='334155')
        
        self.metric_font = Font(name='Arial', size=24, bold=True, color='10B981')
        self.metric_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
        
        self.border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')
    
    def generate_safety_report(
        self,
        report_data: ReportData,
        output_path: str
    ) -> str:
        """Generate comprehensive safety report in Excel"""
        
        self.wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        
        # Create worksheets
        self._create_summary_sheet(report_data)
        self._create_metrics_sheet(report_data.metrics)
        
        if report_data.incidents:
            self._create_incidents_sheet(report_data.incidents)
        
        if report_data.violations:
            self._create_violations_sheet(report_data.violations)
        
        for table in report_data.tables:
            self._create_data_sheet(table)
        
        # Save workbook
        self.wb.save(output_path)
        return output_path
    
    def _create_summary_sheet(self, report_data: ReportData):
        """Create executive summary worksheet"""
        ws = self.wb.create_sheet("Executive Summary", 0)
        
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = report_data.title
        ws[f'A{row}'].font = self.title_font
        ws[f'A{row}'].alignment = self.center_align
        row += 2
        
        # Report Info
        ws[f'A{row}'] = "Organization:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = report_data.organization_name
        row += 1
        
        ws[f'A{row}'] = "Period:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = f"{report_data.period_start.strftime('%Y-%m-%d')} to {report_data.period_end.strftime('%Y-%m-%d')}"
        row += 1
        
        ws[f'A{row}'] = "Generated:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = report_data.generated_at.strftime('%Y-%m-%d %H:%M')
        row += 2
        
        # Executive Summary
        if report_data.executive_summary:
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = "Executive Summary"
            ws[f'A{row}'].font = self.subtitle_font
            row += 1
            
            ws.merge_cells(f'A{row}:F{row+2}')
            ws[f'A{row}'] = report_data.executive_summary
            ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            row += 4
        
        # Key Findings
        if report_data.key_findings:
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = "Key Findings"
            ws[f'A{row}'].font = self.subtitle_font
            row += 1
            
            for finding in report_data.key_findings:
                ws[f'A{row}'] = f"• {finding}"
                ws[f'A{row}'].alignment = self.left_align
                row += 1
            row += 1
        
        # Recommendations
        if report_data.recommendations:
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = "Recommendations"
            ws[f'A{row}'].font = self.subtitle_font
            row += 1
            
            for i, rec in enumerate(report_data.recommendations, 1):
                ws[f'A{row}'] = f"{i}. {rec}"
                ws[f'A{row}'].alignment = self.left_align
                row += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        
    def _create_metrics_sheet(self, metrics: ReportMetrics):
        """Create metrics dashboard worksheet"""
        ws = self.wb.create_sheet("Key Metrics")
        
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "Safety Metrics Dashboard"
        ws[f'A{row}'].font = self.title_font
        ws[f'A{row}'].alignment = self.center_align
        row += 2
        
        # Primary Metrics
        metrics_data = [
            ("Safety Score", metrics.safety_score, "Good" if metrics.safety_score >= 80 else "Needs Improvement"),
            ("Total Incidents", metrics.total_incidents, metrics.trend_direction.capitalize()),
            ("Near Misses", metrics.total_near_misses, ""),
            ("Violations", metrics.total_violations, ""),
            ("Total Alerts", metrics.total_alerts, ""),
            ("Compliance Rate", f"{metrics.compliance_rate:.1f}%", "Good" if metrics.compliance_rate >= 90 else "Action Required"),
            ("PPE Compliance", f"{metrics.ppe_compliance:.1f}%", ""),
            ("Incident Rate", f"{metrics.incident_rate:.2f}", "per 100 workers"),
        ]
        
        # Headers
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Value"
        ws[f'C{row}'] = "Status/Note"
        
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = self.header_font
            ws[f'{col}{row}'].fill = self.header_fill
            ws[f'{col}{row}'].alignment = self.center_align
            ws[f'{col}{row}'].border = self.border
        
        row += 1
        
        # Data rows
        for metric_name, value, status in metrics_data:
            ws[f'A{row}'] = metric_name
            ws[f'B{row}'] = value
            ws[f'C{row}'] = status
            
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            ws[f'C{row}'].border = self.border
            
            ws[f'B{row}'].alignment = self.center_align
            
            # Highlight value cell
            if isinstance(value, (int, float)) and not isinstance(value, str):
                ws[f'B{row}'].font = Font(size=12, bold=True, color='10B981')
            
            row += 1
        
        row += 2
        
        # Critical Incidents
        if metrics.fatalities or metrics.lost_time_incidents:
            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = "Critical Incidents"
            ws[f'A{row}'].font = self.subtitle_font
            row += 1
            
            critical_data = [
                ("Fatalities", metrics.fatalities),
                ("Lost Time Incidents", metrics.lost_time_incidents),
                ("Medical Treatments", metrics.medical_treatments),
                ("First Aid Cases", metrics.first_aid_cases),
            ]
            
            for label, value in critical_data:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                
                # Red highlight for fatalities
                if label == "Fatalities" and value > 0:
                    ws[f'B{row}'].font = Font(size=12, bold=True, color='EF4444')
                
                row += 1
            
            row += 2
        
        # High Risk Areas
        if metrics.high_risk_areas:
            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = "High Risk Areas"
            ws[f'A{row}'].font = self.subtitle_font
            row += 1
            
            for area in metrics.high_risk_areas:
                ws[f'A{row}'] = f"• {area}"
                row += 1
        
        # Auto-adjust columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        
        # Add chart
        self._add_metrics_chart(ws, len(metrics_data))
    
    def _create_incidents_sheet(self, incidents: List[Dict[str, Any]]):
        """Create incidents data worksheet"""
        ws = self.wb.create_sheet("Incidents")
        
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = "Incident Log"
        ws[f'A{row}'].font = self.title_font
        ws[f'A{row}'].alignment = self.center_align
        row += 2
        
        # Headers
        headers = ['ID', 'Date', 'Type', 'Severity', 'Location', 'Status', 'Description']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        
        row += 1
        
        # Data
        for incident in incidents:
            data = [
                incident.get('id', 'N/A'),
                incident.get('date', 'N/A'),
                incident.get('type', 'N/A'),
                incident.get('severity', 'N/A'),
                incident.get('location', 'N/A'),
                incident.get('status', 'N/A'),
                incident.get('description', 'N/A')
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = self.border
                
                # Color code severity
                if col_idx == 4:  # Severity column
                    if value == 'critical':
                        cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                    elif value == 'high':
                        cell.fill = PatternFill(start_color='FED7AA', end_color='FED7AA', fill_type='solid')
            
            row += 1
        
        # Auto-adjust columns
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15
        ws.column_dimensions['G'].width = 40
    
    def _create_violations_sheet(self, violations: List[Dict[str, Any]]):
        """Create violations data worksheet"""
        ws = self.wb.create_sheet("Violations")
        
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "Compliance Violations"
        ws[f'A{row}'].font = self.title_font
        ws[f'A{row}'].alignment = self.center_align
        row += 2
        
        # Headers
        headers = ['ID', 'Date', 'Type', 'Severity', 'Worker', 'Corrective Action']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        
        row += 1
        
        # Data
        for violation in violations:
            data = [
                violation.get('id', 'N/A'),
                violation.get('date', 'N/A'),
                violation.get('type', 'N/A'),
                violation.get('severity', 'N/A'),
                violation.get('worker_id', 'N/A'),
                violation.get('action', 'N/A')
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = self.border
            
            row += 1
        
        # Auto-adjust columns
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _create_data_sheet(self, table: ReportTable):
        """Create custom data worksheet from ReportTable"""
        ws = self.wb.create_sheet(table.title[:31])  # Excel sheet name limit
        
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:{get_column_letter(len(table.headers))}{row}')
        ws[f'A{row}'] = table.title
        ws[f'A{row}'].font = self.title_font
        ws[f'A{row}'].alignment = self.center_align
        row += 2
        
        # Headers
        for col_idx, header in enumerate(table.headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        
        row += 1
        
        # Data rows
        for data_row in table.rows:
            for col_idx, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = self.border
            row += 1
        
        # Totals row
        if table.totals:
            for col_idx, value in enumerate(table.totals, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
                cell.border = self.border
        
        # Auto-adjust columns
        for col in range(1, len(table.headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _add_metrics_chart(self, ws, data_rows: int):
        """Add chart to metrics sheet"""
        chart = BarChart()
        chart.title = "Safety Metrics Overview"
        chart.style = 10
        chart.x_axis.title = "Metrics"
        chart.y_axis.title = "Values"
        
        data = Reference(ws, min_col=2, min_row=4, max_row=4 + data_rows)
        cats = Reference(ws, min_col=1, min_row=5, max_row=4 + data_rows)
        
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "E3")
    
    def generate_quick_export(
        self,
        data: List[Dict[str, Any]],
        headers: List[str],
        output_path: str,
        sheet_name: str = "Data"
    ) -> str:
        """Quick data export to Excel"""
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        
        # Data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
        
        # Auto-adjust columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        wb.save(output_path)
        return output_path
